"""
Serviço de importação de extratos bancários.

Suporta:
- CSV
- Excel (.xlsx)
- PDF (extração de tabelas)

Inclui:
- Detecção automática de colunas
- Suporte multi-moeda (BRL, USD, EUR)
- Detecção de parcelas
- Regras de cartão de crédito (ajuste de data para mês da fatura)
- Detecção de duplicatas
- Validação de saldo
"""

from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import tempfile
import os
import hashlib
import re
import calendar

from fastapi import UploadFile
from sqlalchemy.orm import Session

from app.models import ImportBatch, Transaction, BankAccount, ImportStatus, FileType, CurrencyCode, Category
from app.schemas.import_file import (
    ColumnMapping,
    ImportPreview,
    ImportResult,
    ValidationError,
    DuplicateTransaction
)
from app.services.categorization_service import CategorizationService
from app.services.exchange_service import ExchangeService
from app.services.balance_log_service import log_balance_change


class ColumnDetector:
    """Detecta automaticamente o mapeamento de colunas."""

    DATE_PATTERNS = [
        r'\d{2}/\d{2}/\d{4}',
        r'\d{4}-\d{2}-\d{2}',
        r'\d{2}-\d{2}-\d{4}',
    ]

    DATE_COLUMN_NAMES = [
        'data', 'date', 'dt', 'data_lancamento', 'data_transacao',
        'dt_lancamento', 'dt_transacao', 'movimento', 'data movimento'
    ]

    DESCRIPTION_COLUMN_NAMES = [
        'descricao', 'description', 'desc', 'historico', 'lancamento',
        'memo', 'referencia', 'observacao', 'detalhes', 'descrição'
    ]

    AMOUNT_COLUMN_NAMES = [
        'valor', 'value', 'amount', 'quantia', 'montante',
        'debito', 'credito', 'valor_transacao'
    ]

    # Colunas específicas de moeda
    VALOR_BRL_NAMES = ['valor r$', 'valor_r$', 'valor_brl', 'brl', 'reais', 'valor reais', 'r$', 'valor (r$)', 'valor(r$)']
    VALOR_USD_NAMES = ['valor us$', 'valor_us$', 'valor_usd', 'usd', 'dolar', 'dollar', 'valor dolar', 'us$', 'valor (us$)', 'valor(us$)', 'valor (usd)']
    VALOR_EUR_NAMES = ['valor eur', 'valor_eur', 'eur', 'euro', 'valor euro', 'eur$', 'valor (eur)', 'valor(eur)', '€']

    BALANCE_COLUMN_NAMES = [
        'saldo', 'balance', 'saldo_final', 'saldo_apos',
        'saldo_disponivel', 'saldo_conta'
    ]

    BANK_COLUMN_NAMES = [
        'banco', 'bank', 'instituicao', 'instituição'
    ]

    ACCOUNT_COLUMN_NAMES = [
        'conta', 'account', 'numero_conta', 'conta_numero'
    ]

    CATEGORY_COLUMN_NAMES = [
        'categoria', 'category', 'tipo', 'tipo_despesa', 'classificacao',
        'classificação', 'grupo', 'natureza'
    ]

    CARD_PAYMENT_DATE_NAMES = [
        'data pagto cartao', 'data_pagto_cartao', 'data pagamento',
        'data_pagamento', 'vencimento', 'data vencimento'
    ]

    INSTALLMENT_COLUMN_NAMES = [
        'parcela', 'parcelas', 'installment', 'parc', 'nº parcela',
        'numero parcela', 'numero_parcela'
    ]

    def detect(self, data: List[Dict[str, Any]]) -> ColumnMapping:
        """Detecta mapeamento de colunas automaticamente."""
        if not data:
            raise ValueError("Dados vazios")

        columns = list(data[0].keys())
        sample_rows = data[:10]

        # Detectar cada tipo de coluna
        date_col = self._detect_column(columns, sample_rows, self.DATE_COLUMN_NAMES, 'date')
        desc_col = self._detect_column(columns, sample_rows, self.DESCRIPTION_COLUMN_NAMES, 'text')

        # Detectar colunas de valor (multi-moeda)
        amount_col = self._detect_column(columns, sample_rows, self.AMOUNT_COLUMN_NAMES, 'number')
        valor_brl_col = self._detect_column(columns, sample_rows, self.VALOR_BRL_NAMES, 'number')
        valor_usd_col = self._detect_column(columns, sample_rows, self.VALOR_USD_NAMES, 'number')
        valor_eur_col = self._detect_column(columns, sample_rows, self.VALOR_EUR_NAMES, 'number')

        balance_col = self._detect_column(columns, sample_rows, self.BALANCE_COLUMN_NAMES, 'number')
        bank_col = self._detect_column(columns, sample_rows, self.BANK_COLUMN_NAMES, 'text')
        account_col = self._detect_column(columns, sample_rows, self.ACCOUNT_COLUMN_NAMES, 'text')
        category_col = self._detect_column(columns, sample_rows, self.CATEGORY_COLUMN_NAMES, 'text')
        card_payment_col = self._detect_column(columns, sample_rows, self.CARD_PAYMENT_DATE_NAMES, 'date')
        installment_col = self._detect_column(columns, sample_rows, self.INSTALLMENT_COLUMN_NAMES, 'text')

        return ColumnMapping(
            date_column=date_col or columns[0],
            description_column=desc_col or (columns[1] if len(columns) > 1 else columns[0]),
            amount_column=amount_col or (columns[2] if len(columns) > 2 else None),
            valor_brl_column=valor_brl_col,
            valor_usd_column=valor_usd_col,
            valor_eur_column=valor_eur_col,
            balance_column=balance_col,
            bank_column=bank_col,
            account_column=account_col,
            category_column=category_col,
            card_payment_date_column=card_payment_col,
            installment_column=installment_col
        )

    def _detect_column(
        self,
        columns: List[str],
        sample_rows: List[Dict],
        name_hints: List[str],
        col_type: str
    ) -> Optional[str]:
        """Detecta coluna por nome ou conteúdo."""
        # Por nome (normalizado) - match exato primeiro
        for col in columns:
            col_lower = col.lower().strip()
            if col_lower in name_hints:
                return col
            # Verificar também sem acentos e com underscores
            col_normalized = col_lower.replace(' ', '_').replace('ã', 'a').replace('ç', 'c')
            for hint in name_hints:
                hint_normalized = hint.replace(' ', '_').replace('ã', 'a').replace('ç', 'c')
                if col_normalized == hint_normalized:
                    return col

        # Match parcial (coluna contém o hint ou hint contém a coluna)
        for col in columns:
            col_lower = col.lower().strip()
            for hint in name_hints:
                # Verificar se o nome da coluna contém o hint ou vice-versa
                if hint in col_lower or col_lower in hint:
                    return col

        return None


def detect_installment(description: str) -> Tuple[Optional[int], Optional[int]]:
    """
    Detecta parcela n/m na descrição.

    Args:
        description: Descrição da transação ou conteúdo da coluna parcela

    Returns:
        (parcela_atual, total_parcelas) ou (None, None)
    """
    # Padrão: "7 de 10", "7/10", "03/10", "1 de 12"
    match = re.search(r'(\d{1,2})\s*(?:de|/)\s*(\d{1,2})', description)
    if match:
        n = int(match.group(1))
        m = int(match.group(2))
        # Validar que faz sentido (n <= m e m > 1)
        if 1 <= n <= m and m > 1:
            return n, m
    return None, None


def find_previous_installment(
    db: Session,
    account_id: int,
    description_base: str,
    installment_number: int,
    installment_total: int,
) -> Optional[Transaction]:
    """
    Busca a parcela anterior (n-1) no banco para determinar a data correta.

    Tenta primeiro por campos estruturados (installment_number/total),
    depois por texto na descrição (ex: "6 de 10", "6/10").

    Args:
        db: Sessão do banco
        account_id: ID da conta
        description_base: Descrição sem o sufixo de parcela (ex: "PG *POSITIVO TECN")
        installment_number: Número da parcela atual (ex: 7)
        installment_total: Total de parcelas (ex: 10)

    Returns:
        Transaction da parcela anterior ou None
    """
    if installment_number <= 1:
        return None

    prev_number = installment_number - 1

    # Tentativa 1: campos estruturados
    prev = db.query(Transaction).filter(
        Transaction.account_id == account_id,
        Transaction.installment_number == prev_number,
        Transaction.installment_total == installment_total,
        Transaction.description.ilike(f"%{description_base}%"),
    ).order_by(Transaction.date.desc()).first()

    if prev:
        return prev

    # Tentativa 2: buscar por texto na descrição (ex: "DRASTOSA 6 de 10" ou "DRASTOSA 6/10")
    patterns = [
        f"%{description_base}%{prev_number} de {installment_total}%",
        f"%{description_base}%{prev_number}/{installment_total}%",
        f"%{description_base}%{prev_number:02d}/{installment_total:02d}%",
    ]
    for pattern in patterns:
        prev = db.query(Transaction).filter(
            Transaction.account_id == account_id,
            Transaction.description.ilike(pattern),
        ).order_by(Transaction.date.desc()).first()
        if prev:
            return prev

    return None


def get_installment_base_description(description: str) -> str:
    """
    Remove o sufixo de parcela da descrição.
    Ex: "PG *POSITIVO TECN 6 de 10" -> "PG *POSITIVO TECN"
        "AG DE TURISMO 5/10" -> "AG DE TURISMO"
    """
    # Remove padrões como "7 de 10", "7/10", "07/10"
    cleaned = re.sub(r'\s*\d{1,2}\s*(?:de|/)\s*\d{1,2}\s*$', '', description).strip()
    # Remove traço final se houver (ex: "DRASTOSA -")
    cleaned = re.sub(r'\s*-\s*$', '', cleaned).strip()
    return cleaned


def adjust_date_for_credit_card(original_date: date, payment_date: date) -> date:
    """
    Ajusta data para o mês da fatura do cartão.

    Mantém o dia original se possível, senão usa último dia do mês.

    Args:
        original_date: Data original da transação
        payment_date: Data de pagamento da fatura

    Returns:
        Data ajustada para o mês da fatura
    """
    target_month = payment_date.month
    target_year = payment_date.year

    # Tentar manter o dia
    try:
        return date(target_year, target_month, original_date.day)
    except ValueError:
        # Dia não existe no mês destino (ex: 31 em fevereiro)
        last_day = calendar.monthrange(target_year, target_month)[1]
        return date(target_year, target_month, last_day)


class ImportService:
    """Serviço de importação de extratos bancários."""

    SUPPORTED_TYPES = {
        'text/csv': FileType.CSV,
        'application/vnd.ms-excel': FileType.XLSX,
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': FileType.XLSX,
        'application/pdf': FileType.PDF,
    }

    def __init__(self, db: Session):
        self.db = db
        self.categorization = CategorizationService(db)
        self.exchange_service = ExchangeService(db)
        self.column_detector = ColumnDetector()
        # Cache de categorias por nome para importação histórica
        self._category_cache: Dict[str, int] = {}

    async def upload_and_preview(
        self,
        file: UploadFile,
        account_id: int
    ) -> ImportPreview:
        """Faz upload do arquivo e retorna preview dos dados."""
        # 1. Validar tipo de arquivo
        file_type = self._get_file_type(file)
        if not file_type:
            raise ValueError(f"Tipo de arquivo não suportado: {file.content_type}")

        # 2. Salvar arquivo temporariamente
        temp_path = await self._save_temp_file(file)

        try:
            # 3. Extrair dados usando parser apropriado
            raw_data = self._parse_file(temp_path, file_type)

            if not raw_data:
                raise ValueError("Arquivo vazio ou sem dados válidos")

            # 4. Detectar colunas automaticamente
            column_mapping = self.column_detector.detect(raw_data)

            # 5. Criar batch de importação
            batch = ImportBatch(
                account_id=account_id,
                filename=file.filename,
                file_type=file_type,
                total_records=len(raw_data),
                status=ImportStatus.PENDING,
                temp_file_path=temp_path
            )
            self.db.add(batch)
            self.db.commit()
            self.db.refresh(batch)

            # 6. Preparar preview (primeiras 20 linhas)
            preview_rows = raw_data[:20]

            return ImportPreview(
                batch_id=batch.id,
                total_rows=len(raw_data),
                columns=list(raw_data[0].keys()) if raw_data else [],
                detected_mapping=column_mapping,
                preview_rows=preview_rows,
                temp_file_path=temp_path
            )

        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise

    async def process_import(
        self,
        batch_id: int,
        column_mapping: ColumnMapping,
        account_id: int,
        validate_balance: bool = True,
        expected_final_balance: Optional[Decimal] = None,
        skip_duplicates: bool = True,
        card_payment_date_override: Optional[date] = None
    ) -> ImportResult:
        """Processa a importação com o mapeamento definido."""
        batch = self.db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
        if not batch:
            raise ValueError("Batch não encontrado")

        batch.status = ImportStatus.PROCESSING
        self.db.commit()

        account = self.db.query(BankAccount).filter(BankAccount.id == account_id).first()
        if not account:
            raise ValueError("Conta não encontrada")

        try:
            # 1. Parse do arquivo
            raw_data = self._parse_file(batch.temp_file_path, batch.file_type)

            # 2. Processar cada linha
            transactions = []
            errors = []
            duplicates = []

            # Limpar cache de categorias antes de iniciar
            self._category_cache = {}
            seen_hashes = set()  # Dedup intra-arquivo

            for idx, row in enumerate(raw_data):
                try:
                    result = await self._process_row(
                        row,
                        column_mapping,
                        account_id,
                        batch.id,
                        idx + 1,
                        account,
                        seen_hashes=seen_hashes,
                        card_payment_date_override=card_payment_date_override,
                    )

                    if result['status'] == 'success':
                        transactions.append(result['transaction'])
                    elif result['status'] == 'duplicate':
                        duplicates.append(result['duplicate'])
                    elif result['status'] == 'error':
                        errors.append(result['error'])

                except Exception as e:
                    errors.append(ValidationError(
                        row=idx + 1,
                        field=None,
                        message=str(e)
                    ))

            # 3. Validar saldo se solicitado
            balance_valid = True
            balance_difference = None

            if validate_balance and expected_final_balance is not None:
                # Usar amount_brl para validação de saldo
                total_imported = sum(t.amount_brl for t in transactions)
                calculated_balance = account.current_balance + total_imported
                balance_valid = abs(calculated_balance - expected_final_balance) < Decimal('0.01')
                balance_difference = expected_final_balance - calculated_balance

            # 4. Inserir transações
            if skip_duplicates or not duplicates:
                for t in transactions:
                    self.db.add(t)

                # Atualizar saldo da conta (usar original_amount, que é na moeda da conta)
                total_amount = sum(t.original_amount for t in transactions)
                log_balance_change(self.db, account, account.current_balance + total_amount,
                                   'import', f'Import batch {batch.id}: {len(transactions)} txns')

                batch.imported_records = len(transactions)
                batch.duplicate_records = len(duplicates)
                batch.error_records = len(errors)

                # Preencher período do lote (min/max das datas)
                if transactions:
                    dates = [t.date for t in transactions]
                    batch.date_start = min(dates)
                    batch.date_end = max(dates)

                if duplicates:
                    batch.status = ImportStatus.COMPLETED_WITH_DUPLICATES
                else:
                    batch.status = ImportStatus.COMPLETED
            else:
                batch.status = ImportStatus.FAILED
                batch.error_message = f"Encontradas {len(duplicates)} transações duplicadas"

            self.db.commit()

            # 5. Limpar arquivo temporário
            if batch.temp_file_path and os.path.exists(batch.temp_file_path):
                os.remove(batch.temp_file_path)
                batch.temp_file_path = None
                self.db.commit()

            return ImportResult(
                batch_id=batch.id,
                success=batch.status in [ImportStatus.COMPLETED, ImportStatus.COMPLETED_WITH_DUPLICATES],
                imported_count=len(transactions),
                duplicate_count=len(duplicates),
                error_count=len(errors),
                errors=errors[:10],
                duplicates=duplicates[:10],
                balance_validated=validate_balance,
                balance_matches=balance_valid,
                balance_difference=balance_difference
            )

        except Exception as e:
            self.db.rollback()
            # Recarregar batch após rollback
            batch = self.db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
            if batch:
                batch.status = ImportStatus.FAILED
                batch.error_message = str(e)[:500]  # Limitar tamanho da mensagem
                self.db.commit()
            raise

    async def _process_row(
        self,
        row: Dict[str, Any],
        mapping: ColumnMapping,
        account_id: int,
        batch_id: int,
        row_num: int,
        account: BankAccount,
        seen_hashes: set = None,
        card_payment_date_override: Optional[date] = None,
    ) -> Dict:
        """
        Processa uma linha do arquivo.

        Suporta multi-moeda (BRL, USD, EUR) e detecção de parcelas.
        Para cartões de crédito, ajusta a data para o mês da fatura.
        """
        # Extrair valores usando mapeamento
        date_str = row.get(mapping.date_column)
        description = row.get(mapping.description_column)
        balance_str = row.get(mapping.balance_column) if mapping.balance_column else None
        category_name = row.get(mapping.category_column) if mapping.category_column else None

        # Extrair valores multi-moeda
        amount_str = row.get(mapping.amount_column) if mapping.amount_column else None
        valor_brl_str = row.get(mapping.valor_brl_column) if mapping.valor_brl_column else None
        valor_usd_str = row.get(mapping.valor_usd_column) if mapping.valor_usd_column else None
        valor_eur_str = row.get(mapping.valor_eur_column) if mapping.valor_eur_column else None

        # Data de pagamento do cartão (override global tem prioridade)
        card_payment_str = row.get(mapping.card_payment_date_column) if mapping.card_payment_date_column else None
        # Se há override do formulário, usar ele

        # Validar campos obrigatórios
        if not date_str:
            return {
                'status': 'error',
                'error': ValidationError(row=row_num, field='date', message='Data não informada')
            }

        if not description:
            return {
                'status': 'error',
                'error': ValidationError(row=row_num, field='description', message='Descrição não informada')
            }

        # Pelo menos um valor deve estar preenchido
        has_any_amount = any([
            amount_str and str(amount_str).strip(),
            valor_brl_str and str(valor_brl_str).strip(),
            valor_usd_str and str(valor_usd_str).strip(),
            valor_eur_str and str(valor_eur_str).strip()
        ])

        if not has_any_amount:
            return {
                'status': 'error',
                'error': ValidationError(row=row_num, field='amount', message='Nenhum valor informado')
            }

        # Converter data
        trans_date = self._parse_date(date_str)
        if not trans_date:
            return {
                'status': 'error',
                'error': ValidationError(row=row_num, field='date', message=f'Data inválida: {date_str}')
            }

        # Converter data de pagamento do cartão (se disponível)
        card_payment_date = None
        if card_payment_str:
            card_payment_date = self._parse_date(card_payment_str)
        # Override global do formulário tem prioridade
        if card_payment_date_override:
            card_payment_date = card_payment_date_override

        # Detectar parcelas (coluna dedicada tem prioridade, senão tenta na descrição)
        description_clean = str(description).strip()
        installment_str = row.get(mapping.installment_column) if mapping.installment_column else None

        installment_number, installment_total = None, None
        if installment_str and str(installment_str).strip():
            installment_number, installment_total = detect_installment(str(installment_str))

        if installment_number is None:
            installment_number, installment_total = detect_installment(description_clean)

        # Se parcela veio da coluna dedicada, adicionar ao final da descrição
        if installment_number is not None and installment_str and str(installment_str).strip():
            # Só adicionar se a descrição ainda não contém info de parcela
            existing_n, _ = detect_installment(description_clean)
            if existing_n is None:
                description_clean = f"{description_clean} {installment_number} de {installment_total}"

        # Converter valores multi-moeda
        valor_brl = self._parse_amount(valor_brl_str) if valor_brl_str and str(valor_brl_str).strip() else None
        valor_usd = self._parse_amount(valor_usd_str) if valor_usd_str and str(valor_usd_str).strip() else None
        valor_eur = self._parse_amount(valor_eur_str) if valor_eur_str and str(valor_eur_str).strip() else None

        # Parse amount genérico
        amount_generic = self._parse_amount(amount_str) if amount_str and str(amount_str).strip() else None

        # Se só temos amount genérico e não temos valores específicos, usar como valor da moeda da conta
        if amount_generic is not None and not any([valor_brl, valor_usd, valor_eur]):
            if account.currency == CurrencyCode.BRL:
                valor_brl = amount_generic
            elif account.currency == CurrencyCode.USD:
                valor_usd = amount_generic
            elif account.currency == CurrencyCode.EUR:
                valor_eur = amount_generic
            else:
                # Default: assume BRL
                valor_brl = amount_generic

        # Determinar moeda original e valor original
        original_currency = account.currency
        original_amount = None

        # Prioridade: moeda da conta > BRL > USD > EUR
        if account.currency == CurrencyCode.BRL and valor_brl is not None:
            original_currency = CurrencyCode.BRL
            original_amount = valor_brl
        elif account.currency == CurrencyCode.USD and valor_usd is not None:
            original_currency = CurrencyCode.USD
            original_amount = valor_usd
        elif account.currency == CurrencyCode.EUR and valor_eur is not None:
            original_currency = CurrencyCode.EUR
            original_amount = valor_eur
        elif valor_brl is not None:
            original_currency = CurrencyCode.BRL
            original_amount = valor_brl
        elif valor_usd is not None:
            original_currency = CurrencyCode.USD
            original_amount = valor_usd
        elif valor_eur is not None:
            original_currency = CurrencyCode.EUR
            original_amount = valor_eur

        if original_amount is None:
            return {
                'status': 'error',
                'error': ValidationError(row=row_num, field='amount', message='Valor inválido - nenhum valor encontrado')
            }

        # Calcular valores convertidos (usar valores do CSV se disponíveis, senão converter)
        amount_brl = valor_brl
        amount_usd = valor_usd
        amount_eur = valor_eur

        # Garantir que pelo menos amount_brl existe (para compatibilidade)
        if amount_brl is None and original_currency == CurrencyCode.BRL:
            amount_brl = original_amount

        # Converter valores faltantes usando serviço de câmbio
        try:
            if amount_brl is None:
                if original_currency == CurrencyCode.BRL:
                    amount_brl = original_amount
                else:
                    amount_brl = await self.exchange_service.convert(
                        original_amount, original_currency, CurrencyCode.BRL, trans_date
                    )

            if amount_usd is None:
                if original_currency == CurrencyCode.USD:
                    amount_usd = original_amount
                else:
                    try:
                        amount_usd = await self.exchange_service.convert(
                            original_amount, original_currency, CurrencyCode.USD, trans_date
                        )
                    except ValueError:
                        amount_usd = Decimal("0.00")

            if amount_eur is None:
                if original_currency == CurrencyCode.EUR:
                    amount_eur = original_amount
                else:
                    try:
                        amount_eur = await self.exchange_service.convert(
                            original_amount, original_currency, CurrencyCode.EUR, trans_date
                        )
                    except ValueError:
                        amount_eur = Decimal("0.00")
        except ValueError as e:
            # Se não conseguir converter para BRL, é erro crítico
            if amount_brl is None:
                amount_brl = original_amount if original_currency == CurrencyCode.BRL else Decimal("0.00")
            if amount_usd is None:
                amount_usd = Decimal("0.00")
            if amount_eur is None:
                amount_eur = Decimal("0.00")

        # Converter saldo (se disponível)
        balance_after = None
        if balance_str:
            balance_after = self._parse_amount(balance_str)

        # Ajustar data para cartão de crédito com parcelas
        effective_date = trans_date
        if account.is_credit_card and card_payment_date and installment_number is not None:
            # Tentar encontrar a parcela anterior no banco para calcular data correta
            desc_base = get_installment_base_description(description_clean)
            prev_installment = find_previous_installment(
                self.db, account_id, desc_base, installment_number, installment_total
            )

            if prev_installment:
                # Usar data da parcela anterior + 1 mês
                prev_date = prev_installment.date
                next_month = prev_date.month + 1
                next_year = prev_date.year
                if next_month > 12:
                    next_month = 1
                    next_year += 1
                try:
                    effective_date = date(next_year, next_month, prev_date.day)
                except ValueError:
                    # Dia não existe no mês (ex: 31 em mês de 30 dias)
                    last_day = calendar.monthrange(next_year, next_month)[1]
                    effective_date = date(next_year, next_month, last_day)
            else:
                # Fallback: usar lógica original (ajustar para mês da fatura)
                effective_date = adjust_date_for_credit_card(trans_date, card_payment_date)

        # Gerar hash para verificar duplicata (normalização é feita dentro de generate_hash)
        trans_hash = Transaction.generate_hash(
            account_id,
            effective_date,
            description_clean,
            original_amount,
            original_currency,
            card_payment_date=card_payment_date
        )

        # === Detecção de duplicados multi-camada ===

        # Layer 1: Hash exato (indexado, O(1))
        is_duplicate = False
        existing = None

        # Para duplicatas intra-arquivo (mesmo hash dentro do mesmo CSV):
        # Não tratar como duplicata — podem ser transações legítimas distintas
        # (ex: "PG *POSITIVO TECNO06/10" 2x R$413,62 no mesmo extrato)
        # Em vez disso, adicionar sufixo para tornar o hash único
        if seen_hashes is not None and trans_hash in seen_hashes:
            suffix = 1
            suffixed_hash = Transaction.generate_hash(
                account_id, effective_date, description_clean,
                original_amount, original_currency, suffix,
                card_payment_date=card_payment_date
            )
            while suffixed_hash in seen_hashes:
                suffix += 1
                suffixed_hash = Transaction.generate_hash(
                    account_id, effective_date, description_clean,
                    original_amount, original_currency, suffix,
                    card_payment_date=card_payment_date
                )
            # Verificar se o hash com sufixo já existe no banco
            existing_suffixed = self.db.query(Transaction).filter(
                Transaction.transaction_hash == suffixed_hash
            ).first()
            if existing_suffixed:
                is_duplicate = True
                existing = existing_suffixed
            else:
                trans_hash = suffixed_hash

        if not is_duplicate:
            existing = self.db.query(Transaction).filter(
                Transaction.transaction_hash == trans_hash
            ).first()
            if existing:
                is_duplicate = True

        # Layer 2: Fuzzy matching (se Layer 1 não encontrou)
        if not is_duplicate:
            from app.utils.normalization import find_near_duplicate
            fuzzy_match = find_near_duplicate(
                self.db, account_id, effective_date,
                original_amount, original_currency, description_clean,
                card_payment_date=card_payment_date
            )
            if fuzzy_match:
                existing = fuzzy_match
                is_duplicate = True

        if is_duplicate:
            return {
                'status': 'duplicate',
                'duplicate': DuplicateTransaction(
                    row=row_num,
                    date=effective_date.isoformat(),
                    description=description_clean,
                    amount=amount_brl,
                    existing_id=existing.id if existing else 0
                )
            }

        # Registrar hash no set intra-arquivo
        if seen_hashes is not None:
            seen_hashes.add(trans_hash)

        # Tentar categorizar automaticamente ou usar categoria do CSV
        category_id = None
        is_validated = False

        if category_name and str(category_name).strip():
            # Buscar categoria por nome (case insensitive)
            category_id = self._get_category_id_by_name(str(category_name).strip())
            if category_id:
                is_validated = True

        if not category_id:
            # Tentar categorização automática
            category_id, confidence, method = self.categorization.categorize(
                description_clean,
                float(amount_brl)
            )
            is_validated = (method == 'rule')

        # Aprender com a categorização (CSV ou automática) para melhorar futuras importações
        if category_id:
            self.categorization.learn_from_categorization(description_clean, category_id)

        # Criar transação
        transaction = Transaction(
            account_id=account_id,
            category_id=category_id,
            date=effective_date,
            description=description_clean,
            original_description=description_clean,
            original_currency=original_currency,
            original_amount=original_amount,
            amount_brl=amount_brl,
            amount_usd=amount_usd,
            amount_eur=amount_eur,
            amount=original_amount,  # Valor na moeda da conta
            balance_after=balance_after,
            card_payment_date=card_payment_date,
            installment_number=installment_number,
            installment_total=installment_total,
            transaction_hash=trans_hash,
            is_validated=is_validated,
            import_batch_id=batch_id
        )

        return {'status': 'success', 'transaction': transaction}

    def _get_category_id_by_name(self, name: str) -> Optional[int]:
        """
        Busca ID da categoria pelo nome (case insensitive).
        Usa cache para evitar queries repetidas.
        """
        name_lower = name.lower().strip()

        if name_lower in self._category_cache:
            return self._category_cache[name_lower]

        category = self.db.query(Category).filter(
            Category.name.ilike(name_lower)
        ).first()

        if category:
            self._category_cache[name_lower] = category.id
            return category.id

        return None

    def _parse_date(self, value: Any) -> Optional[date]:
        """Converte string para data."""
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        formats = [
            '%d/%m/%Y',
            '%Y-%m-%d',
            '%d-%m-%Y',
            '%d.%m.%Y',
            '%d/%m/%y',
            '%Y/%m/%d',
        ]

        str_value = str(value).strip()
        for fmt in formats:
            try:
                return datetime.strptime(str_value, fmt).date()
            except ValueError:
                continue

        return None

    def _parse_amount(self, value: Any) -> Optional[Decimal]:
        """Converte string para valor decimal."""
        if isinstance(value, (int, float)):
            return Decimal(str(value))

        if value is None:
            return None

        str_value = str(value).strip()

        # Remover símbolos de moeda
        str_value = str_value.replace('R$', '').replace('$', '').strip()

        # Detectar formato brasileiro (1.234,56) vs americano (1,234.56)
        if ',' in str_value and '.' in str_value:
            if str_value.rfind(',') > str_value.rfind('.'):
                # Formato brasileiro
                str_value = str_value.replace('.', '').replace(',', '.')
            else:
                # Formato americano
                str_value = str_value.replace(',', '')
        elif ',' in str_value:
            # Assumir brasileiro se só tem vírgula
            str_value = str_value.replace(',', '.')

        # Remover espaços
        str_value = str_value.replace(' ', '')

        try:
            return Decimal(str_value)
        except InvalidOperation:
            return None

    def _get_file_type(self, file: UploadFile) -> Optional[FileType]:
        """Detecta tipo do arquivo."""
        # Por content type
        if file.content_type in self.SUPPORTED_TYPES:
            return self.SUPPORTED_TYPES[file.content_type]

        # Por extensão
        if file.filename:
            ext = file.filename.lower().split('.')[-1]
            if ext == 'csv':
                return FileType.CSV
            elif ext in ('xlsx', 'xls'):
                return FileType.XLSX
            elif ext == 'pdf':
                return FileType.PDF

        return None

    async def _save_temp_file(self, file: UploadFile) -> str:
        """Salva arquivo em diretório temporário."""
        suffix = '.' + file.filename.split('.')[-1] if file.filename else ''

        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
            content = await file.read()
            temp.write(content)
            return temp.name

    def _parse_file(self, file_path: str, file_type: FileType) -> List[Dict[str, Any]]:
        """Parse do arquivo baseado no tipo."""
        if file_type == FileType.CSV:
            return self._parse_csv(file_path)
        elif file_type == FileType.XLSX:
            return self._parse_excel(file_path)
        elif file_type == FileType.PDF:
            return self._parse_pdf(file_path)
        else:
            raise ValueError(f"Tipo de arquivo não suportado: {file_type}")

    def _parse_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse arquivo CSV."""
        import csv

        # Detectar encoding (utf-8-sig strips BOM automatically)
        encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        content = None

        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue

        if content is None:
            raise ValueError("Não foi possível detectar encoding do arquivo")

        # Detectar delimitador
        delimiters = [',', ';', '\t', '|']
        delimiter = max(delimiters, key=lambda d: content.count(d))

        # Ler CSV
        rows = []
        with open(file_path, 'r', encoding=enc) as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            for row in reader:
                # Limpar tanto as chaves (nomes de coluna) quanto os valores
                cleaned = {k.strip() if k else k: v.strip() if v else '' for k, v in row.items()}
                rows.append(cleaned)

        return rows

    def _parse_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse arquivo Excel (.xlsx ou .xls)."""
        import os
        ext = os.path.splitext(file_path)[1].lower()

        if ext == '.xls':
            return self._parse_xls(file_path)

        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return []

        # Primeira linha como cabeçalho
        headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]

        # Converter para dicionários
        result = []
        for row in rows[1:]:
            if any(cell is not None for cell in row):
                row_dict = {
                    headers[i]: row[i] if i < len(row) else None
                    for i in range(len(headers))
                }
                result.append(row_dict)

        wb.close()
        return result

    def _parse_xls(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse arquivo Excel .xls (formato antigo)."""
        import xlrd

        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        if ws.nrows == 0:
            return []

        # Primeira linha como cabeçalho
        headers = [
            str(ws.cell_value(0, j)).strip() if ws.cell_value(0, j) else f'col_{j}'
            for j in range(ws.ncols)
        ]

        # Converter para dicionários
        result = []
        for i in range(1, ws.nrows):
            row_vals = [ws.cell_value(i, j) for j in range(ws.ncols)]
            if any(v is not None and v != '' for v in row_vals):
                row_dict = {
                    headers[j]: row_vals[j] if row_vals[j] != '' else None
                    for j in range(len(headers))
                }
                result.append(row_dict)

        return result

    def _parse_pdf(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse PDF usando tabula."""
        try:
            import tabula
            dfs = tabula.read_pdf(file_path, pages='all', multiple_tables=True)

            if not dfs:
                raise ValueError("Nenhuma tabela encontrada no PDF")

            # Concatenar todas as tabelas
            all_rows = []
            for df in dfs:
                records = df.to_dict('records')
                all_rows.extend(records)

            return all_rows
        except ImportError:
            raise ValueError("Biblioteca tabula-py não instalada para processamento de PDF")
