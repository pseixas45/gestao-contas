"""Parser da Posição Detalhada Histórica da XP em formato .xlsx.

O arquivo `XP_PosicaoDetalhadaHistorica_DD_MM_AAAA.xlsx` tem uma aba
("Sua carteira") com um resumo no topo e seções por subcategoria. Cada seção
começa com uma linha de cabeçalho no formato:

    ['13,9% | Inflação', 'Posição', '% Alocação', ..., 'Valor líquido']

seguida das linhas de posição alinhadas a esse cabeçalho. Há dois layouts
(fundos e renda fixa), tratados de forma genérica lendo os nomes das colunas
do próprio cabeçalho de cada seção.
"""
import re
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional

from app.services.parsers.base import (
    ParsedPosition, ParsedSnapshot,
    parse_money, parse_pct, parse_date_br, normalize_name, detect_asset_class,
)

FILENAME_DATE_RE = re.compile(r"_(\d{2})_(\d{2})_(\d{4})")


class XPXlsxParser:
    """Parser da Posição Detalhada Histórica da XP (.xlsx)."""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)

    def parse(self) -> ParsedSnapshot:
        from openpyxl import load_workbook

        snapshot_date = self._extract_date_from_filename()
        if not snapshot_date:
            raise ValueError(f"Não foi possível extrair data do nome: {self.file_path.name}")

        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb["Sua carteira"] if "Sua carteira" in wb.sheetnames else wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

        available_balance = Decimal("0")
        positions: List[ParsedPosition] = []
        current_class = None
        col_map: Dict[str, int] = {}

        for r in rows:
            if not r:
                continue
            first = str(r[0]).strip() if r[0] is not None else ""
            if not first:
                continue

            # Resumo do topo: capturar "Saldo Disponível histórico"
            if "Saldo Disponível" in " ".join(str(c) for c in r if c):
                # a linha seguinte (valores) é tratada normalmente; ignora header
                pass
            low = first.lower()
            if low.startswith("saldo disponível") or low.startswith("saldo disponivel"):
                v = parse_money(r[1]) if len(r) > 1 else None
                if v is not None:
                    available_balance = v
                continue

            # Cabeçalho de seção: "X% | Subcategoria" + nomes de colunas
            if "|" in first and detect_asset_class(first):
                current_class = detect_asset_class(first)
                col_map = {}
                for idx, cell in enumerate(r[1:], start=1):
                    if cell is None:
                        continue
                    name = str(cell).strip().lower()
                    if name:
                        col_map[name] = idx
                continue

            # Linhas de subtotal/agrupamento (ex: "Renda Fixa", "Fundos de Investimentos")
            # têm poucas colunas e não têm classe corrente aplicável a elas.
            if current_class is None:
                continue

            # Linha de posição: precisa ter um valor de mercado na coluna "Posição"
            value = self._col(r, col_map, ("posição a mercado", "posição", "posicao a mercado", "posicao"))
            if value is None:
                continue

            pos: ParsedPosition = {
                "name": first,
                "name_normalized": normalize_name(first),
                "asset_class": current_class,
                "value": value,
            }
            invested = self._col(r, col_map, ("valor aplicado",), exclude=("original",))
            if invested is not None:
                pos["value_invested"] = invested
            net = self._col(r, col_map, ("valor líquido", "valor liquido"))
            if net is not None:
                pos["value_net"] = net
            alloc = self._col_pct(r, col_map, ("% alocação", "% alocacao"))
            if alloc is not None:
                pos["allocation_pct"] = alloc
            qty = self._col(r, col_map, ("quantidade",))
            if qty is not None:
                pos["quantity"] = qty
            mat = self._col_date(r, col_map, ("data vencimento",))
            if mat is not None:
                pos["maturity_date"] = mat
            app_date = self._col_date(r, col_map, ("data aplicação", "data aplicacao"))
            if app_date is not None:
                pos["application_date"] = app_date
            rate = self._col_raw(r, col_map, ("taxa a mercado",))
            if rate:
                pos["contracted_rate"] = rate
            positions.append(pos)

        total_value = sum((p["value"] for p in positions), Decimal("0"))

        return {
            "snapshot_date": snapshot_date,
            "total_value": total_value,
            "total_invested": sum(
                (p.get("value_invested") or Decimal("0") for p in positions), Decimal("0")
            ) or None,
            "available_balance": available_balance,
            "positions": positions,
        }

    # ---- helpers ----

    def _extract_date_from_filename(self) -> Optional[date]:
        m = FILENAME_DATE_RE.search(self.file_path.name)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return date(y, mo, d)
            except ValueError:
                return None
        return None

    def _idx(self, col_map: Dict[str, int], names, exclude=()):
        for key, idx in col_map.items():
            if any(n in key for n in names) and not any(e in key for e in exclude):
                return idx
        return None

    def _col_raw(self, r, col_map, names, exclude=()):
        idx = self._idx(col_map, names, exclude)
        if idx is None or idx >= len(r):
            return None
        return r[idx]

    def _col(self, r, col_map, names, exclude=()):
        return parse_money(self._col_raw(r, col_map, names, exclude))

    def _col_pct(self, r, col_map, names):
        return parse_pct(self._col_raw(r, col_map, names))

    def _col_date(self, r, col_map, names):
        return parse_date_br(self._col_raw(r, col_map, names))
