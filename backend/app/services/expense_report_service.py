"""
Serviço para gestão de relatórios de reembolso de despesas de trabalho.
"""

import re
from io import BytesIO
from decimal import Decimal
from typing import List, Optional
from datetime import datetime
from collections import defaultdict

from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from fastapi import HTTPException

from app.models.transaction import Transaction
from app.models.account import BankAccount
from app.models.category import Category
from app.models.expense_report import ExpenseReport, ExpenseReportItem, ExpenseReportStatus
from app.schemas.expense_report import (
    ExpenseReportCreate,
    ExpenseReportUpdate,
    ExpenseReportDetail,
    ExpenseReportSummary,
    ExpenseReportTransactionItem,
    UnreportedTransaction,
    ExpectedItem,
    ExpectedItemsResponse,
)


WORK_EXPENSE_CATEGORY_NAME = "Despesas Trabalho"


class ExpenseReportService:
    def __init__(self, db: Session):
        self.db = db

    def _get_work_expense_category_id(self) -> int:
        cat = self.db.query(Category).filter(Category.name == WORK_EXPENSE_CATEGORY_NAME).first()
        if not cat:
            raise HTTPException(status_code=404, detail=f"Categoria '{WORK_EXPENSE_CATEGORY_NAME}' não encontrada")
        return cat.id

    def get_unreported_transactions(self) -> List[UnreportedTransaction]:
        category_id = self._get_work_expense_category_id()

        reported_ids = self.db.query(ExpenseReportItem.transaction_id).subquery()

        transactions = (
            self.db.query(Transaction, BankAccount.name.label("account_name"))
            .join(BankAccount, Transaction.account_id == BankAccount.id)
            .filter(
                Transaction.category_id == category_id,
                ~Transaction.id.in_(self.db.query(reported_ids.c.transaction_id)),
            )
            .order_by(Transaction.date.desc())
            .all()
        )

        result = []
        for t, account_name in transactions:
            result.append(UnreportedTransaction(
                id=t.id,
                date=t.date,
                description=t.description,
                amount_brl=abs(t.amount_brl),
                original_amount=abs(t.original_amount),
                original_currency=t.original_currency.value if t.original_currency else "BRL",
                installment_info=t.installment_info if t.is_installment else None,
                account_name=account_name,
            ))
        return result

    def create_report(self, data: ExpenseReportCreate) -> ExpenseReportDetail:
        category_id = self._get_work_expense_category_id()

        # Validar que transações existem e pertencem à categoria
        transactions = (
            self.db.query(Transaction)
            .filter(Transaction.id.in_(data.transaction_ids))
            .all()
        )

        if len(transactions) != len(data.transaction_ids):
            found_ids = {t.id for t in transactions}
            missing = set(data.transaction_ids) - found_ids
            raise HTTPException(status_code=400, detail=f"Transações não encontradas: {missing}")

        wrong_category = [t.id for t in transactions if t.category_id != category_id]
        if wrong_category:
            raise HTTPException(
                status_code=400,
                detail=f"Transações não pertencem à categoria '{WORK_EXPENSE_CATEGORY_NAME}': {wrong_category}",
            )

        # Verificar se alguma já está em outro relatório
        already_reported = (
            self.db.query(ExpenseReportItem.transaction_id)
            .filter(ExpenseReportItem.transaction_id.in_(data.transaction_ids))
            .all()
        )
        if already_reported:
            ids = [r.transaction_id for r in already_reported]
            raise HTTPException(status_code=400, detail=f"Transações já incluídas em outro relatório: {ids}")

        # Calcular total
        total_brl = sum(abs(t.amount_brl) for t in transactions)

        # Criar relatório
        report = ExpenseReport(
            reference_month=data.reference_month,
            status=ExpenseReportStatus.draft,
            notes=data.notes,
            total_brl=total_brl,
        )
        self.db.add(report)
        self.db.flush()

        # Criar itens
        for t in transactions:
            item = ExpenseReportItem(report_id=report.id, transaction_id=t.id)
            self.db.add(item)

        self.db.commit()
        self.db.refresh(report)

        return self.get_report(report.id)

    def get_report(self, report_id: int) -> ExpenseReportDetail:
        report = (
            self.db.query(ExpenseReport)
            .options(joinedload(ExpenseReport.items).joinedload(ExpenseReportItem.transaction))
            .filter(ExpenseReport.id == report_id)
            .first()
        )
        if not report:
            raise HTTPException(status_code=404, detail="Relatório não encontrado")

        # Buscar nomes das contas
        account_ids = {item.transaction.account_id for item in report.items}
        accounts = {a.id: a.name for a in self.db.query(BankAccount).filter(BankAccount.id.in_(account_ids)).all()}

        items = []
        for item in sorted(report.items, key=lambda x: x.transaction.date):
            t = item.transaction
            items.append(ExpenseReportTransactionItem(
                transaction_id=t.id,
                date=t.date,
                description=t.description,
                amount_brl=abs(t.amount_brl),
                original_amount=abs(t.original_amount),
                original_currency=t.original_currency.value if t.original_currency else "BRL",
                installment_info=t.installment_info if t.is_installment else None,
                account_name=accounts.get(t.account_id),
            ))

        return ExpenseReportDetail(
            id=report.id,
            reference_month=report.reference_month,
            status=report.status.value,
            total_brl=report.total_brl,
            item_count=len(report.items),
            notes=report.notes,
            created_at=report.created_at,
            updated_at=report.updated_at,
            items=items,
        )

    def list_reports(self) -> List[ExpenseReportSummary]:
        # Query com count via subquery para evitar N+1
        item_count_sq = (
            self.db.query(
                ExpenseReportItem.report_id,
                func.count(ExpenseReportItem.id).label("cnt"),
            )
            .group_by(ExpenseReportItem.report_id)
            .subquery()
        )

        rows = (
            self.db.query(ExpenseReport, func.coalesce(item_count_sq.c.cnt, 0).label("item_count"))
            .outerjoin(item_count_sq, ExpenseReport.id == item_count_sq.c.report_id)
            .order_by(ExpenseReport.reference_month.desc())
            .all()
        )

        return [
            ExpenseReportSummary(
                id=r.id,
                reference_month=r.reference_month,
                status=r.status.value,
                total_brl=r.total_brl,
                item_count=cnt,
                notes=r.notes,
                created_at=r.created_at,
                updated_at=r.updated_at,
            )
            for r, cnt in rows
        ]

    def update_report(self, report_id: int, data: ExpenseReportUpdate) -> ExpenseReportDetail:
        report = self.db.query(ExpenseReport).filter(ExpenseReport.id == report_id).first()
        if not report:
            raise HTTPException(status_code=404, detail="Relatório não encontrado")

        # Atualizar status
        if data.status:
            new_status = ExpenseReportStatus(data.status)
            valid_transitions = {
                ExpenseReportStatus.draft: {ExpenseReportStatus.submitted},
                ExpenseReportStatus.submitted: {ExpenseReportStatus.reimbursed, ExpenseReportStatus.draft},
                ExpenseReportStatus.reimbursed: set(),
            }
            if new_status not in valid_transitions.get(report.status, set()):
                raise HTTPException(
                    status_code=400,
                    detail=f"Transição de status inválida: {report.status.value} → {data.status}",
                )
            report.status = new_status

        # Modificar itens só se rascunho
        if data.add_transaction_ids or data.remove_transaction_ids:
            if report.status != ExpenseReportStatus.draft:
                raise HTTPException(status_code=400, detail="Só é possível modificar itens em relatórios rascunho")

        if data.remove_transaction_ids:
            self.db.query(ExpenseReportItem).filter(
                ExpenseReportItem.report_id == report_id,
                ExpenseReportItem.transaction_id.in_(data.remove_transaction_ids),
            ).delete(synchronize_session=False)

        if data.add_transaction_ids:
            category_id = self._get_work_expense_category_id()
            # Validar transações
            transactions = (
                self.db.query(Transaction)
                .filter(Transaction.id.in_(data.add_transaction_ids))
                .all()
            )
            wrong_category = [t.id for t in transactions if t.category_id != category_id]
            if wrong_category:
                raise HTTPException(status_code=400, detail=f"Transações fora da categoria: {wrong_category}")

            already_reported = (
                self.db.query(ExpenseReportItem.transaction_id)
                .filter(ExpenseReportItem.transaction_id.in_(data.add_transaction_ids))
                .all()
            )
            if already_reported:
                ids = [r.transaction_id for r in already_reported]
                raise HTTPException(status_code=400, detail=f"Transações já em outro relatório: {ids}")

            for t in transactions:
                self.db.add(ExpenseReportItem(report_id=report_id, transaction_id=t.id))

        if data.notes is not None:
            report.notes = data.notes

        # Recalcular total
        if data.add_transaction_ids or data.remove_transaction_ids:
            self.db.flush()
            items = self.db.query(ExpenseReportItem).filter(ExpenseReportItem.report_id == report_id).all()
            if items:
                t_ids = [i.transaction_id for i in items]
                total = self.db.query(func.sum(func.abs(Transaction.amount_brl))).filter(
                    Transaction.id.in_(t_ids)
                ).scalar()
                report.total_brl = total or Decimal("0.00")
            else:
                report.total_brl = Decimal("0.00")

        self.db.commit()
        return self.get_report(report_id)

    def delete_report(self, report_id: int) -> None:
        report = self.db.query(ExpenseReport).filter(ExpenseReport.id == report_id).first()
        if not report:
            raise HTTPException(status_code=404, detail="Relatório não encontrado")
        if report.status != ExpenseReportStatus.draft:
            raise HTTPException(status_code=400, detail="Só é possível excluir relatórios rascunho")

        self.db.delete(report)
        self.db.commit()

    @staticmethod
    def _normalize_description(desc: str) -> str:
        """Normaliza descrição para agrupar variações do mesmo item."""
        s = desc.upper().strip()
        s = re.sub(r'\d{2}/\d{2}', '', s)          # remove dd/mm (parcelas cartão)
        s = re.sub(r'\d+ DE \d+', '', s)            # remove "1 de 3"
        s = re.sub(r'PARCELA \d+', '', s)           # remove "parcela N"
        s = re.sub(r'\d{2,}', '', s)                # remove sequências numéricas longas
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    def get_expected_items(self, lookback: int = 6) -> ExpectedItemsResponse:
        """Analisa últimos N relatórios e identifica itens recorrentes.
        Cruza com transações não reportadas para verificar quais estão presentes."""

        # Buscar últimos N relatórios (excluir drafts)
        reports = (
            self.db.query(ExpenseReport)
            .filter(ExpenseReport.status != ExpenseReportStatus.draft)
            .order_by(ExpenseReport.reference_month.desc())
            .limit(lookback)
            .all()
        )

        if not reports:
            return ExpectedItemsResponse(expected=[], found_count=0, missing_count=0)

        total_reports = len(reports)

        # Coletar descrições normalizadas por relatório
        pattern_months = defaultdict(set)       # pattern -> set of months
        pattern_amounts = defaultdict(list)     # pattern -> list of amounts
        pattern_sample = {}                     # pattern -> most recent description

        for report in reports:
            items = (
                self.db.query(ExpenseReportItem)
                .options(joinedload(ExpenseReportItem.transaction))
                .filter(ExpenseReportItem.report_id == report.id)
                .all()
            )
            for item in items:
                t = item.transaction
                norm = self._normalize_description(t.description)
                pattern_months[norm].add(report.reference_month)
                pattern_amounts[norm].append(abs(t.amount_brl))
                # Manter sample mais recente
                if norm not in pattern_sample or report.reference_month > pattern_sample[norm][1]:
                    pattern_sample[norm] = (t.description, report.reference_month)

        # Filtrar: mínimo 50% de frequência (ou pelo menos 3 ocorrências)
        min_freq = max(2, total_reports // 2)
        recurring = {
            p: len(months)
            for p, months in pattern_months.items()
            if len(months) >= min_freq
        }

        if not recurring:
            return ExpectedItemsResponse(expected=[], found_count=0, missing_count=0)

        # Buscar transações não reportadas
        unreported = self.get_unreported_transactions()

        # Criar lookup de não-reportadas por pattern
        unreported_by_pattern = defaultdict(list)
        for t in unreported:
            norm = self._normalize_description(t.description)
            unreported_by_pattern[norm].append(t)

        # Montar resultado
        expected = []
        for pattern, freq in sorted(recurring.items(), key=lambda x: -x[1]):
            amounts = pattern_amounts[pattern]
            avg_amount = Decimal(str(round(sum(amounts) / len(amounts), 2)))
            matches = unreported_by_pattern.get(pattern, [])

            expected.append(ExpectedItem(
                pattern=pattern,
                sample_description=pattern_sample[pattern][0],
                frequency=freq,
                total_reports=total_reports,
                avg_amount=avg_amount,
                found=len(matches) > 0,
                matched_transaction_ids=[m.id for m in matches],
            ))

        found = sum(1 for e in expected if e.found)
        missing = sum(1 for e in expected if not e.found)

        return ExpectedItemsResponse(
            expected=expected,
            found_count=found,
            missing_count=missing,
        )

    def export_to_excel(self, report_id: int) -> tuple[BytesIO, str]:
        """Gera planilha Excel do relatório. Retorna (buffer, filename)."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, numbers

        detail = self.get_report(report_id)

        # Mês por extenso
        month_names = {
            "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
            "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
            "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro",
        }
        year, month = detail.reference_month.split("-")
        month_label = f"{month_names.get(month, month)}/{year}"

        status_labels = {"draft": "Rascunho", "submitted": "Enviado", "reimbursed": "Reembolsado"}

        wb = Workbook()
        ws = wb.active
        ws.title = "Reembolso"

        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        total_font = Font(bold=True, size=11)
        currency_fmt = '#,##0.00'

        # Cabeçalho do relatório
        ws.merge_cells("A1:F1")
        ws["A1"] = "Relatório de Reembolso - Despesas Trabalho"
        ws["A1"].font = title_font

        ws.merge_cells("A2:F2")
        ws["A2"] = f"Mês de Referência: {month_label}"

        ws.merge_cells("A3:F3")
        ws["A3"] = f"Status: {status_labels.get(detail.status, detail.status)} | Gerado em: {datetime.now().strftime('%d/%m/%Y')}"

        # Colunas do cabeçalho
        headers = ["Data", "Descrição", "Valor (R$)", "Moeda Original", "Valor Original", "Parcela"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Dados
        for row_idx, item in enumerate(detail.items, 6):
            ws.cell(row=row_idx, column=1, value=item.date.strftime("%d/%m/%Y"))
            ws.cell(row=row_idx, column=2, value=item.description)
            cell_brl = ws.cell(row=row_idx, column=3, value=float(item.amount_brl))
            cell_brl.number_format = currency_fmt
            ws.cell(row=row_idx, column=4, value=item.original_currency)
            cell_orig = ws.cell(row=row_idx, column=5, value=float(item.original_amount))
            cell_orig.number_format = currency_fmt
            ws.cell(row=row_idx, column=6, value=item.installment_info or "")

        # Linha total
        total_row = 6 + len(detail.items) + 1
        ws.cell(row=total_row, column=2, value="TOTAL").font = total_font
        total_cell = ws.cell(row=total_row, column=3, value=float(detail.total_brl))
        total_cell.font = total_font
        total_cell.number_format = currency_fmt

        # Largura das colunas
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 10

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"reembolso_{detail.reference_month}.xlsx"
        return buffer, filename
