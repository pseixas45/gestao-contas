"""Importa Itau 260501.xlsx filtrando apenas range 06/03 a 30/04/2026.

1. Cria xlsx temporario com linhas apenas do range desejado
2. Faz upload via API
3. Processa com skip_duplicates
4. Valida saldo final = R$ 1.066,66
"""
import openpyxl
import requests
import sys
from datetime import date, datetime
from pathlib import Path
import tempfile

API_URL = "http://localhost:8002"
USERNAME = "pseixas"
PASSWORD = "pseixas123"
SOURCE = r'c:\Users\paulo\gestao-contas\extratos\Itau 260501.xlsx'
ACCOUNT_ID = 5  # Itaú
CUTOFF_MIN = date(2026, 3, 6)
CUTOFF_MAX = date(2026, 4, 30)
EXPECTED_BALANCE = 1066.66


def parse_dt(s):
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    try:
        return datetime.strptime(str(s).strip(), '%d/%m/%Y').date()
    except (ValueError, TypeError):
        return None


def filter_file():
    """Cria arquivo xlsx temporário com apenas linhas no range desejado."""
    wb_in = openpyxl.load_workbook(SOURCE, read_only=True)
    sh_in = wb_in.active

    wb_out = openpyxl.Workbook()
    sh_out = wb_out.active

    rows = list(sh_in.iter_rows(values_only=True))
    # Header (preservar primeiras 2 linhas: column names)
    if rows:
        sh_out.append(rows[0])  # data, lançamento, ag./origem, valor (R$), saldos (R$)
        sh_out.append(rows[1])  # lancamentos / vazio

    kept = 0
    skipped_pre = 0
    skipped_future = 0
    skipped_saldo = 0

    for row in rows[2:]:
        if not row or len(row) < 4:
            continue
        d = parse_dt(row[0])
        if not d:
            continue
        # Linha de saldo - preservar para validação interna do importador
        if row[1] and 'SALDO' in str(row[1]).upper():
            # Manter linhas de saldo do range para o sistema poder validar
            if CUTOFF_MIN <= d <= CUTOFF_MAX:
                sh_out.append(row)
                continue
            skipped_saldo += 1
            continue
        # Filtrar por data
        if d < CUTOFF_MIN:
            skipped_pre += 1
            continue
        if d > CUTOFF_MAX:
            skipped_future += 1
            continue
        sh_out.append(row)
        kept += 1

    # Salvar em temp
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_path = tmp.name
    tmp.close()
    wb_out.save(tmp_path)
    print(f"Arquivo filtrado: {tmp_path}")
    print(f"  Linhas mantidas: {kept}")
    print(f"  Skipped pre-06/03: {skipped_pre}")
    print(f"  Skipped futuras (>30/04): {skipped_future}")
    print(f"  Skipped saldos fora do range: {skipped_saldo}")
    return tmp_path


def login():
    r = requests.post(
        f"{API_URL}/api/v1/auth/login",
        data={"username": USERNAME, "password": PASSWORD},
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )
    r.raise_for_status()
    return r.json()["access_token"]


def import_file(token, file_path, custom_name='Itau 260501 filtered.xlsx'):
    headers = {"Authorization": f"Bearer {token}"}

    # Upload
    with open(file_path, "rb") as f:
        upload_resp = requests.post(
            f"{API_URL}/api/v1/imports/upload",
            files={"file": (custom_name, f)},
            data={"account_id": ACCOUNT_ID},
            headers=headers,
        )
    upload_resp.raise_for_status()
    upload_data = upload_resp.json()
    batch_id = upload_data["batch_id"]
    mapping = upload_data.get("detected_mapping")
    print(f"\nUpload OK: batch_id={batch_id}")
    print(f"  Mapeamento detectado: {mapping}")

    if not mapping:
        print(f"ERRO: nenhum mapeamento detectado")
        return None

    # Process
    print(f"\nProcessando...")
    process_resp = requests.post(
        f"{API_URL}/api/v1/imports/process",
        json={
            "batch_id": batch_id,
            "account_id": ACCOUNT_ID,
            "column_mapping": mapping,
            "validate_balance": False,
            "skip_duplicates": True,
        },
        headers=headers,
    )
    process_resp.raise_for_status()
    pdata = process_resp.json()
    print(f"  Sucesso: {pdata.get('success')}")
    print(f"  Importadas: {pdata.get('imported_count')}")
    print(f"  Duplicatas: {pdata.get('duplicate_count')}")
    print(f"  Erros: {pdata.get('error_count')}")
    if pdata.get('errors'):
        print(f"\n  Detalhe dos erros:")
        for err in pdata.get('errors', [])[:10]:
            print(f"    {err}")
    return pdata


def validate_balance(token):
    """Verifica current_balance final"""
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.get(f"{API_URL}/api/v1/accounts/{ACCOUNT_ID}", headers=headers)
    r.raise_for_status()
    bal = float(r.json()['current_balance'])
    diff = bal - EXPECTED_BALANCE
    print(f"\n=== VALIDACAO FINAL ===")
    print(f"  current_balance: R$ {bal:,.2f}")
    print(f"  esperado: R$ {EXPECTED_BALANCE:,.2f}")
    print(f"  diferenca: R$ {diff:,.2f}")
    if abs(diff) < 0.01:
        print(f"  >>> SALDO BATEU! <<<")
        return True
    print(f"  >>> ATENCAO: divergencia detectada")
    return False


def main():
    print("=== Filtragem do arquivo ===")
    tmp_path = filter_file()

    print(f"\n=== Login ===")
    token = login()
    print("OK")

    print(f"\n=== Importacao ===")
    result = import_file(token, tmp_path)
    if result is None:
        return 1

    ok = validate_balance(token)

    # Cleanup
    Path(tmp_path).unlink(missing_ok=True)

    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
