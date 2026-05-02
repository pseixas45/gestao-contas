"""Importa C6 R$ pré-processando o arquivo.

C6 R$ usa duas colunas separadas (Entrada/Saída) que o ImportService
não consolida automaticamente. Pré-processamos criando coluna 'Valor'
unificada e enviamos esse arquivo modificado.
"""
import openpyxl
import requests
import os
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path

API = 'http://localhost:8002'
SOURCE = r'c:\Users\paulo\gestao-contas\extratos\C6 R$ 260501.xlsx'
ACCOUNT_ID = 3
EXPECTED_BALANCE = 208.28


def parse_dt(s):
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    try:
        return datetime.strptime(str(s).strip(), '%d/%m/%Y').date()
    except (ValueError, TypeError):
        return None


def transform_file():
    """Cria xlsx com coluna 'Valor' (Entrada - Saída) e descrição combinada."""
    wb_in = openpyxl.load_workbook(SOURCE, read_only=True)
    sh_in = wb_in.active

    wb_out = openpyxl.Workbook()
    sh_out = wb_out.active

    # Header novo: Data, Descricao, Valor, Saldo
    sh_out.append(['Data', 'Descricao', 'Valor', 'Saldo'])

    rows = list(sh_in.iter_rows(values_only=True))
    kept = 0
    for row in rows[1:]:
        if not row or len(row) < 6:
            continue
        d = parse_dt(row[0])
        if not d:
            continue
        titulo = row[2] or ''
        sub = row[3] or ''
        # Combinar titulo + descricao para preservar detalhes
        desc = f"{titulo} ({sub})" if sub and sub != titulo else str(titulo)
        try:
            entrada = float(row[4] or 0)
            saida = float(row[5] or 0)
        except (ValueError, TypeError):
            continue
        valor = entrada - saida
        if valor == 0:
            continue
        saldo = row[6] if len(row) > 6 else None
        sh_out.append([d.strftime('%d/%m/%Y'), desc, valor, saldo])
        kept += 1

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_path = tmp.name
    tmp.close()
    wb_out.save(tmp_path)
    print(f'Arquivo transformado: {tmp_path}')
    print(f'  Linhas mantidas: {kept}')
    return tmp_path


def main():
    print('=== Transformando arquivo ===')
    tmp = transform_file()

    print('\n=== Login ===')
    token = requests.post(
        f'{API}/api/v1/auth/login',
        data={'username': 'pseixas', 'password': 'pseixas123'},
        headers={'Content-Type': 'application/x-www-form-urlencoded'}
    ).json()['access_token']
    H = {'Authorization': f'Bearer {token}'}
    print('OK')

    print('\n=== Upload ===')
    with open(tmp, 'rb') as f:
        r = requests.post(
            f'{API}/api/v1/imports/upload',
            files={'file': ('C6Rs_260501_processado.xlsx', f)},
            data={'account_id': ACCOUNT_ID},
            headers=H,
        )
    r.raise_for_status()
    data = r.json()
    batch_id = data['batch_id']
    mapping = data.get('detected_mapping')
    print(f'  batch_id: {batch_id}')
    print(f'  mapping: {mapping}')

    if not mapping:
        print('ERRO: sem mapping')
        return 1

    print('\n=== Processando ===')
    r = requests.post(
        f'{API}/api/v1/imports/process',
        json={
            'batch_id': batch_id,
            'account_id': ACCOUNT_ID,
            'column_mapping': mapping,
            'validate_balance': False,
            'skip_duplicates': True,
        },
        headers=H,
    )
    r.raise_for_status()
    res = r.json()
    print(f'  importadas: {res.get("imported_count")}')
    print(f'  duplicatas: {res.get("duplicate_count")}')
    print(f'  erros: {res.get("error_count")}')

    # Validar
    r = requests.get(f'{API}/api/v1/accounts/{ACCOUNT_ID}', headers=H)
    bal = float(r.json()['current_balance'])
    print(f'\n=== VALIDACAO ===')
    print(f'  current_balance: R$ {bal:,.2f}')
    print(f'  esperado: R$ {EXPECTED_BALANCE:,.2f}')
    print(f'  diff: R$ {bal - EXPECTED_BALANCE:,.2f}')
    ok = abs(bal - EXPECTED_BALANCE) < 0.01
    print(f'  {"BATEU!" if ok else "DIVERGENCIA"}')

    Path(tmp).unlink(missing_ok=True)
    return 0 if ok else 1


if __name__ == '__main__':
    sys.exit(main())
