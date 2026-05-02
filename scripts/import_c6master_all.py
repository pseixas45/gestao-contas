"""Importa os 3 arquivos C6 Master (cartão C6 Carbon).

Pré-processa cada arquivo:
- Remove linhas extras de cabeçalho (ex: "Nome: PAULO ESTEVAM SEIXAS...")
- Mantém apenas o header das colunas + transações
- Importa via API com card_payment_date correto
"""
import openpyxl
import requests
import os
import sys
import tempfile
from pathlib import Path

API = 'http://localhost:8002'
ACCOUNT_ID = 2  # C6M


def transform(source_path):
    """Cria xlsx limpo só com header + transações."""
    wb_in = openpyxl.load_workbook(source_path, read_only=True)
    sh_in = wb_in.active

    rows = list(sh_in.iter_rows(values_only=True))
    # Detectar linha do header REAL (a que tem 'Data de compra')
    header_idx = None
    for i, row in enumerate(rows):
        if row and any(c and 'Data de compra' in str(c) for c in row):
            header_idx = i
            break
    if header_idx is None:
        print(f'ERRO: header não encontrado em {source_path}')
        return None

    wb_out = openpyxl.Workbook()
    sh_out = wb_out.active
    for row in rows[header_idx:]:
        sh_out.append(row)

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_path = tmp.name
    tmp.close()
    wb_out.save(tmp_path)
    return tmp_path


def main():
    token = requests.post(
        f'{API}/api/v1/auth/login',
        data={'username': 'pseixas', 'password': 'pseixas123'},
        headers={'Content-Type': 'application/x-www-form-urlencoded'}
    ).json()['access_token']
    H = {'Authorization': f'Bearer {token}'}

    requests.delete(f'{API}/api/v1/imports/templates/{ACCOUNT_ID}', headers=H)

    bal = float(requests.get(f'{API}/api/v1/accounts/{ACCOUNT_ID}', headers=H).json()['current_balance'])
    print(f'Saldo antes: R$ {bal:,.2f}')

    files = [
        (r'c:\Users\paulo\gestao-contas\extratos\C6 Master 260310.xlsx', '2026-03-10'),
        (r'c:\Users\paulo\gestao-contas\extratos\C6 Master 260410.xlsx', '2026-04-10'),
        (r'c:\Users\paulo\gestao-contas\extratos\C6 Master 260510.xlsx', '2026-05-10'),
    ]

    for path, cpd in files:
        fname = os.path.basename(path)
        print(f'\n== {fname} (cpd={cpd}) ==')
        tmp = transform(path)
        if not tmp:
            continue

        with open(tmp, 'rb') as f:
            r = requests.post(f'{API}/api/v1/imports/upload',
                              files={'file': (fname, f)},
                              data={'account_id': ACCOUNT_ID}, headers=H)
        data = r.json()
        batch_id = data['batch_id']
        cols = data['columns']

        # Mapping correto: amount_column = 'Valor (em R$)' (nao Cotacao nem US$)
        mapping = dict(data['detected_mapping'])
        valor_brl = None
        for c in cols:
            cl = c.lower().strip()
            if cl.startswith('valor') and 'r' in cl and 'us' not in cl and 'eur' not in cl:
                valor_brl = c
                break
        if valor_brl:
            mapping['valor_brl_column'] = valor_brl
            mapping['amount_column'] = None
            mapping['valor_usd_column'] = None
        print(f'  valor_brl_column = {valor_brl}')
        print(f'  description_column = {mapping.get("description_column")}')
        print(f'  installment_column = {mapping.get("installment_column")}')

        r = requests.post(f'{API}/api/v1/imports/process', json={
            'batch_id': batch_id, 'account_id': ACCOUNT_ID, 'column_mapping': mapping,
            'validate_balance': False, 'skip_duplicates': True,
            'card_payment_date': cpd,
        }, headers=H)
        res = r.json()
        print(f'  imp={res.get("imported_count")} dup={res.get("duplicate_count")} err={res.get("error_count")}')
        if res.get('errors'):
            for e in res['errors'][:3]:
                print(f'    err: {e}')
        bal = float(requests.get(f'{API}/api/v1/accounts/{ACCOUNT_ID}', headers=H).json()['current_balance'])
        print(f'  saldo: R$ {bal:,.2f}')

        Path(tmp).unlink(missing_ok=True)

    # Validacao
    print()
    print(f'== VALIDACAO FINAL ==')
    print(f'  saldo: R$ {bal:,.2f}')
    print(f'  esperado: R$ -22,054.96')
    print(f'  diff: R$ {bal - (-22054.96):,.2f}')
    print(f'  {"BATEU!" if abs(bal - (-22054.96)) < 0.01 else "DIVERGENCIA"}')


if __name__ == '__main__':
    sys.exit(main() or 0)
