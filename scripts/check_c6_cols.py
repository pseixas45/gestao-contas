import openpyxl
import os

base = r'c:\Users\paulo\gestao-contas\extratos'
for f in ['C6 Master 260310.xlsx', 'C6 Master 260410.xlsx', 'C6 Master 260510.xlsx']:
    path = os.path.join(base, f)
    wb = openpyxl.load_workbook(path, read_only=True)
    sh = wb.active
    print(f'== {f} ==')
    for i, row in enumerate(sh.iter_rows(values_only=True)):
        if i == 0:
            for j, c in enumerate(row):
                print(f'  col{j}: {repr(c)}')
            break
    # Linha 1 (primeira transação)
    rows = list(sh.iter_rows(values_only=True))
    if len(rows) > 1:
        print(f'  Linha 1 (transacao): {rows[1]}')
    print()
